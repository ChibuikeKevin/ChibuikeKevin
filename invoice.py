from openpyxl import Workbook

def generate_invoice_number(counter):
    return str(counter).zfill(6)

wb = Workbook()
ws = wb.active
ws['F4'] = 'invoice_number'

row = 5
counter = 1
while True:
    invoice_number = generate_invoice_number(counter)
    ws[f'F{row}'] = invoice_number
    print(invoice_number)
    row += 1
    counter += 1
    if row > 50:  # Save the file every 50 rows to avoid data loss
        wb.save('invoices_number.xlsx')
        row = 5  # Reset row to overwrite the file from the beginning